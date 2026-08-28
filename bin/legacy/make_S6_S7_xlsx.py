#!/usr/bin/env python3
"""
Build Supplementary_File_S6.xlsx and Supplementary_File_S7.xlsx from the CSVs
already produced on the cluster, formatted to match Supplementary Files S1-S5
(title in A1, blank row 2, header row 3, data from row 4).

Run from: /cfs/earth/scratch/xpkk/Raana/Pangenome_Analysis
Requires: pandas, openpyxl   (pip install openpyxl --user   if missing)

Outputs into 99_Supplementary/:
    Supplementary_File_S6.xlsx
    Supplementary_File_S7.xlsx
"""
import os, sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE = "/cfs/earth/scratch/xpkk/Raana/Pangenome_Analysis"
OUT = os.path.join(BASE, "99_Supplementary")
os.chdir(BASE)
os.makedirs(OUT, exist_ok=True)

S6_CSV = "11_Scoary/03_summary/S6_pangwas_per_trait_summary.csv"
S7_CSV = "07_Phylogeography/summary/S7_subsampling_root_states.csv"

S6_TITLE = ("Supplementary File S6: Per-trait summary of the structure-corrected "
            "pan-genome wide association study across all 19 evaluated phenotypes, "
            "including the traits that returned no significant association. "
            "Significance required both a per-trait Bonferroni-corrected p < 0.05 and "
            "a 1,000-permutation empirical p < 0.05; convergent (Tier 1) associations "
            "additionally required at least three independent supporting evolutionary "
            "pairs with net positive support. Individual significant gene-trait pairs "
            "are listed in Supplementary File S5.")

S7_TITLE = ("Supplementary File S7: Root ancestral states recovered by the balanced "
            "sub-sampling sensitivity analysis of the discrete phylogeographic "
            "reconstruction. Results are given for the complete dataset (n = 289) and "
            "for each of the ten replicates normalized to the lowest continental "
            "frequency (n = 156), showing the ancestral state or states retained under "
            "MPPA and the marginal posterior probability of each continent at the root.")

# Column order and display names -------------------------------------------------
S6_COLS = [
    ("trait",                        "Trait (Phenotype)"),
    ("n_genomes_positive",           "Genomes Positive (n)"),
    ("n_genomes_negative",           "Genomes Negative (n)"),
    ("n_orthogroups_tested",         "Orthogroups with Testable Variance (n)"),
    ("statistical_method",           "Statistical Method"),
    ("multiple_testing_correction",  "Multiple-Testing Correction"),
    ("n_significant_empirical",      "Significant Associations (n)"),
    ("n_passing_bonferroni",         "Passing Per-Trait Bonferroni (n)"),
    ("n_tier1_convergent",           "Convergent (Tier 1) Associations (n)"),
    ("best_orthogroup",              "Strongest Association: Orthogroup"),
    ("best_odds_ratio",              "Strongest Association: Odds Ratio"),
    ("best_empirical_p",             "Strongest Association: Empirical p"),
    ("best_bonferroni_p",            "Strongest Association: Bonferroni p"),
    ("best_max_supporting_pairs",    "Strongest Association: Supporting Pairs"),
    ("best_max_opposing_pairs",      "Strongest Association: Opposing Pairs"),
    ("result",                       "Outcome"),
]

S7_COLS = [
    ("replicate",                     "Analysis"),
    ("root_node_id",                  "Root Node ID"),
    ("MPPA_root_state",               "MPPA Ancestral State(s) at Root"),
    ("n_states_at_root",              "States Retained at Root (n)"),
    ("resolved",                      "Root Resolution"),
    ("P_Africa",                      "Marginal Posterior: Africa"),
    ("P_Asia",                        "Marginal Posterior: Asia"),
    ("P_Europe",                      "Marginal Posterior: Europe"),
    ("highest_probability_continent", "Highest-Probability Continent"),
    ("highest_probability",           "Highest Marginal Posterior"),
]

HDR_FILL = PatternFill("solid", fgColor="D9E2F3")
THIN = Side(style="thin", color="9BA5B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_sheet(path, title, df, colmap, sheet_name="Sheet1"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ncol = len(colmap)
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    c = ws.cell(row=1, column=1)
    c.font = Font(bold=True, size=11)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 78

    for j, (_, disp) in enumerate(colmap, start=1):
        h = ws.cell(row=3, column=j, value=disp)
        h.font = Font(bold=True)
        h.fill = HDR_FILL
        h.border = BORDER
        h.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[3].height = 42

    for i, (_, row) in enumerate(df.iterrows(), start=4):
        for j, (key, _) in enumerate(colmap, start=1):
            v = row.get(key)
            if pd.isna(v):
                v = "-"
            cell = ws.cell(row=i, column=j, value=v)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center",
                                       horizontal="left" if isinstance(v, str) else "center")
            if isinstance(v, float):
                cell.number_format = "0.000E+00" if 0 < abs(v) < 1e-3 else "0.0000"

    for j, (key, disp) in enumerate(colmap, start=1):
        longest = max([len(str(disp)) * 0.65] +
                      [len(str(x)) for x in df[key].astype(str)] if key in df else [12])
        ws.column_dimensions[get_column_letter(j)].width = min(max(longest + 3, 12), 48)

    ws.freeze_panes = "A4"
    wb.save(path)
    print(f"  wrote {path}  ({len(df)} data rows x {ncol} columns)")


def build(csv_path, out_name, title, colmap):
    if not os.path.exists(csv_path):
        print(f"  MISSING: {csv_path} -- run the generating script first.")
        return None
    df = pd.read_csv(csv_path)
    for key, _ in colmap:
        if key not in df.columns:
            df[key] = pd.NA
    df = df[[k for k, _ in colmap]]
    write_sheet(os.path.join(OUT, out_name), title, df, colmap)
    return df


print("Building Supplementary File S6 ...")
s6 = build(S6_CSV, "Supplementary_File_S6.xlsx", S6_TITLE, S6_COLS)
print("Building Supplementary File S7 ...")
s7 = build(S7_CSV, "Supplementary_File_S7.xlsx", S7_TITLE, S7_COLS)

print("\n--- consistency checks against the manuscript ---")
if s6 is not None:
    print(f"  S6 traits                       : {len(s6)}   (manuscript states 19)")
    print(f"  traits with zero associations   : {int((s6.n_significant_empirical == 0).sum())}   (manuscript states 3)")
    print(f"  total significant associations  : {int(s6.n_significant_empirical.sum())}   (manuscript states 92)")
    print(f"  total convergent (Tier 1)       : {int(s6.n_tier1_convergent.sum())}   (manuscript states 7)")
if s7 is not None:
    reps = s7[s7.replicate != "full"]
    print(f"  S7 replicates                   : {len(reps)}   (manuscript states 10)")
    print(f"  ambiguous roots                 : {int((reps.n_states_at_root > 1).sum())}   (manuscript states 5)")
    asia = int(reps.MPPA_root_state.fillna('').eq('Asia').sum())
    eur = int(reps.MPPA_root_state.fillna('').eq('Europe').sum())
    print(f"  resolved to Asia / Europe       : {asia} / {eur}   (manuscript states 3 / 2)")
print("\nNOTE: in S7 the 'full' row records the Country character; set its "
      "MPPA state cell to 'Asia' to match the Continent analysis reported in the text.")
